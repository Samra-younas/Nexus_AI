from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.exceptions import RequestValidationError
from pydantic import BaseModel
from typing import List, Optional
import anthropic
import os
import base64
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

app = FastAPI()

# ── Anthropic client ──────────────────────────────────────────────────────────
client = anthropic.Anthropic(
    api_key=os.getenv("ANTHROPIC_API_KEY"),
    base_url=os.getenv("ANTHROPIC_BASE_URL")
)

MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")

SYSTEM_PROMPT = """You are Claude, an AI assistant made by Anthropic.
Never say you are Cascade or any other AI.
Answer questions thoroughly and accurately.
Use markdown formatting, bullet points, and structure when it helps clarity.
If a user asks you to generate a PDF, Word document, or any downloadable file,
explain clearly that this app does not support file generation/download yet,
and instead offer to provide the full content as formatted text they can copy."""

# ── Supabase client ───────────────────────────────────────────────────────────
try:
    from supabase import create_client, Client
    _sb_url = os.getenv("SUPABASE_URL", "")
    _sb_key = os.getenv("SUPABASE_KEY", "")
    supabase: Optional[Client] = create_client(_sb_url, _sb_key) if _sb_url and _sb_key else None
except ImportError:
    supabase = None
    print("WARNING: supabase-py not installed. Run: pip install supabase")

# ── File type constants ───────────────────────────────────────────────────────
IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}

TEXT_EXTENSIONS = {
    "py", "js", "ts", "jsx", "tsx", "html", "css", "scss",
    "json", "xml", "yaml", "yml", "toml", "ini", "cfg", "env",
    "txt", "md", "markdown", "csv", "sql", "sh", "bash",
    "c", "cpp", "h", "java", "go", "rs", "php", "rb", "swift",
    "kt", "dart", "r", "m", "tf", "dockerfile", "makefile",
    "gitignore", "lock", "log"
}

TEXT_MIME_TYPES = {
    "application/json", "application/xml", "application/javascript",
    "application/typescript", "application/x-python", "application/x-sh"
}

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Global error handlers ─────────────────────────────────────────────────────

@app.exception_handler(RequestValidationError)
async def validation_error_handler(request: Request, exc: RequestValidationError):
    return JSONResponse(status_code=422, content={"error": f"Invalid request: {exc.errors()}"})

@app.exception_handler(Exception)
async def global_error_handler(request: Request, exc: Exception):
    return JSONResponse(status_code=500, content={"error": f"Server error: {str(exc)}"})


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_text_file(filename: str, mime: str) -> bool:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in TEXT_EXTENSIONS: return True
    if mime.startswith("text/"): return True
    if mime in TEXT_MIME_TYPES: return True
    return False


def extract_docx_text(file_bytes: bytes) -> str:
    try:
        import docx
        doc = docx.Document(BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return "[python-docx not installed — run: pip install python-docx]"
    except Exception as e:
        return f"[Could not extract .docx content: {e}]"


def extract_xlsx_text(file_bytes: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(c) if c is not None else "" for c in row)
                if row_text.strip():
                    lines.append(row_text)
        return "\n".join(lines)
    except ImportError:
        return "[openpyxl not installed — run: pip install openpyxl]"
    except Exception as e:
        return f"[Could not extract .xlsx content: {e}]"


def get_text_from_response(response) -> str:
    text_parts = []
    has_tool_use = False
    for block in response.content:
        if hasattr(block, "text") and block.text:
            text_parts.append(block.text)
        elif hasattr(block, "type") and block.type == "tool_use":
            has_tool_use = True
    if text_parts:
        return "\n".join(text_parts)
    if has_tool_use:
        return (
            "⚠️ I tried to use a tool to complete that request (like generating a file), "
            "but this app doesn't support file downloads yet.\n\n"
            "**What I can do instead:** Ask me to write the full content as text — "
            "you can copy it, or I can format it as markdown."
        )
    return "I processed your request but had no text response to return. Please try again."


# ── Pydantic models ───────────────────────────────────────────────────────────

class Message(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    messages: List[Message]

class SessionSave(BaseModel):
    session_id: str
    name: Optional[str] = None
    messages: list        # display messages (JSON)
    api_messages: list    # api history (JSON)

class SessionRename(BaseModel):
    session_id: str
    name: str


# ── Chat routes ───────────────────────────────────────────────────────────────

@app.post("/chat")
def chat(req: ChatRequest):
    try:
        valid_messages = [
            {"role": m.role, "content": m.content}
            for m in req.messages
            if m.role in ("user", "assistant") and m.content.strip()
        ]
        if not valid_messages:
            return {"error": "No valid messages to send."}
        response = client.messages.create(
            model=MODEL, max_tokens=4096,
            system=SYSTEM_PROMPT, messages=valid_messages
        )
        return {"answer": get_text_from_response(response)}
    except anthropic.APIStatusError as e:
        return {"error": f"API error {e.status_code}: {e.message}"}
    except anthropic.APIConnectionError:
        return {"error": "Could not connect to AI service. Check your API key and base URL."}
    except Exception as e:
        return {"error": f"Unexpected error: {str(e)}"}


@app.post("/chat-files")
async def chat_files(message: str = Form(""), files: List[UploadFile] = File(...)):
    try:
        content_blocks = []

        for f in files:
            try:
                file_bytes = await f.read()
            except Exception as e:
                content_blocks.append({"type": "text", "text": f"[Could not read file '{f.filename}': {e}]"})
                continue

            filename = f.filename or "file"
            mime = (f.content_type or "application/octet-stream").split(";")[0].strip()
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

            if mime in IMAGE_TYPES:
                try:
                    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
                    content_blocks.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
                    content_blocks.append({"type": "text", "text": f"[Image: {filename}]"})
                except Exception as e:
                    content_blocks.append({"type": "text", "text": f"[Failed to encode image '{filename}': {e}]"})

            elif mime == "application/pdf" or ext == "pdf":
                try:
                    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
                    content_blocks.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
                    content_blocks.append({"type": "text", "text": f"[PDF: {filename}]"})
                except Exception as e:
                    content_blocks.append({"type": "text", "text": f"[Failed to process PDF '{filename}': {e}]"})

            elif mime == DOCX_MIME or ext == "docx":
                content_blocks.append({"type": "text", "text": f"[Word Document: {filename}]\n{extract_docx_text(file_bytes)}"})

            elif mime == XLSX_MIME or ext == "xlsx":
                content_blocks.append({"type": "text", "text": f"[Excel Spreadsheet: {filename}]\n{extract_xlsx_text(file_bytes)}"})

            elif is_text_file(filename, mime):
                try:
                    text_content = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    text_content = file_bytes.decode("latin-1", errors="replace")
                content_blocks.append({"type": "text", "text": f"[File: {filename}]\n```{ext}\n{text_content}\n```"})

            else:
                size_kb = round(len(file_bytes) / 1024, 1)
                content_blocks.append({"type": "text", "text": f"[Binary file: {filename} ({mime}, {size_kb} KB) — cannot read binary content directly]"})

        user_text = message.strip() if message.strip() else "Please analyze the attached file(s)."
        content_blocks.append({"type": "text", "text": user_text})

        response = client.messages.create(
            model=MODEL, max_tokens=4096,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content_blocks}]
        )
        return {"answer": get_text_from_response(response)}

    except anthropic.APIStatusError as e:
        return {"error": f"API error {e.status_code}: {e.message}"}
    except anthropic.APIConnectionError:
        return {"error": "Could not connect to AI service. Check your API key and base URL."}
    except Exception as e:
        return {"error": f"Unexpected error: {str(e)}"}


# ── Supabase session routes ───────────────────────────────────────────────────
# Table schema (run once in Supabase SQL editor):
#
# create table sessions (
#   session_id text primary key,
#   name text,
#   messages jsonb not null default '[]',
#   api_messages jsonb not null default '[]',
#   created_at timestamptz default now(),
#   updated_at timestamptz default now()
# );

@app.get("/sessions")
def get_sessions():
    """Load all sessions (sidebar list)."""
    if not supabase:
        return {"error": "Supabase not configured.", "sessions": []}
    try:
        res = supabase.table("sessions") \
            .select("session_id, name, created_at, updated_at") \
            .order("updated_at", desc=True) \
            .execute()
        return {"sessions": res.data}
    except Exception as e:
        return {"error": str(e), "sessions": []}


@app.get("/sessions/{session_id}")
def get_session(session_id: str):
    """Load a single session's full messages."""
    if not supabase:
        return {"error": "Supabase not configured."}
    try:
        res = supabase.table("sessions") \
            .select("*") \
            .eq("session_id", session_id) \
            .single() \
            .execute()
        return {"session": res.data}
    except Exception as e:
        return {"error": str(e)}


@app.post("/sessions")
def save_session(data: SessionSave):
    """Create or update a session (upsert)."""
    if not supabase:
        return {"error": "Supabase not configured."}
    try:
        supabase.table("sessions").upsert({
            "session_id": data.session_id,
            "name": data.name,
            "messages": data.messages,
            "api_messages": data.api_messages,
            "updated_at": "now()"
        }).execute()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


@app.patch("/sessions/{session_id}/rename")
def rename_session(session_id: str, data: SessionRename):
    """Rename a session."""
    if not supabase:
        return {"error": "Supabase not configured."}
    try:
        supabase.table("sessions") \
            .update({"name": data.name, "updated_at": "now()"}) \
            .eq("session_id", session_id) \
            .execute()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


@app.delete("/sessions/{session_id}")
def delete_session(session_id: str):
    """Delete a session."""
    if not supabase:
        return {"error": "Supabase not configured."}
    try:
        supabase.table("sessions") \
            .delete() \
            .eq("session_id", session_id) \
            .execute()
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


# ── Static files ──────────────────────────────────────────────────────────────

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("static/index.html")